import openpyxl

source_path = "/Users/kjw1/Desktop/Monitoring/332044_proj3_popr.xlsx"
reference_path = "/Users/kjw1/Desktop/Monitoring/Proj3/Reference.xlsx"

for label, path, sheet in [
    ("SOURCE", source_path, "Przemieszczenia ostateczne"),
    ("REFERENCE", reference_path, "Przemieszczenia "),
]:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[sheet]
    print(f"\n--- {label} {sheet} formulas ---")
    for r in range(1, 56):
        values = [ws.cell(r, c).value for c in range(1, 18)]
        if any(v is not None for v in values):
            print(r, values)

for label, path, sheet in [
    ("SOURCE", source_path, "Przemieszczenia ostateczne"),
    ("REFERENCE", reference_path, "Przemieszczenia "),
]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    print(f"\n--- {label} {sheet} cached values ---")
    for r in range(1, 56):
        values = [ws.cell(r, c).value for c in range(1, 18)]
        if any(v is not None for v in values):
            print(r, values)
