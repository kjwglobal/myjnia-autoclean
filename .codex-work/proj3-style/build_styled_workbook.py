from __future__ import annotations

import re
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SOURCE_PATH = Path("/Users/kjw1/Desktop/Monitoring/332044_proj3_popr.xlsx")
REFERENCE_PATH = Path("/Users/kjw1/Desktop/Monitoring/Proj3/Reference.xlsx")
OUTPUT_DIR = Path("/Users/kjw1/Documents/New project/outputs/proj3_style_20260612")
OUTPUT_PATH = OUTPUT_DIR / "332044_proj3_popr_styl_reference.xlsx"


def copy_cell(src, dst, copy_value: bool) -> None:
    if copy_value:
        dst.value = src.value
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def copy_dimensions(src_ws, dst_ws, max_row: int, max_col: int) -> None:
    for row_idx in range(1, max_row + 1):
        src_dim = src_ws.row_dimensions[row_idx]
        dst_dim = dst_ws.row_dimensions[row_idx]
        if src_dim.height is not None:
            dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        src_dim = src_ws.column_dimensions[letter]
        dst_dim = dst_ws.column_dimensions[letter]
        if src_dim.width is not None:
            dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden


def clear_merges(ws) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))


def copy_merges(src_ws, dst_ws, max_row: int, max_col: int) -> None:
    for merged_range in src_ws.merged_cells.ranges:
        if merged_range.max_row <= max_row and merged_range.max_col <= max_col:
            dst_ws.merge_cells(str(merged_range))


def copy_sheet_area(src_ws, dst_ws, max_row: int, max_col: int, copy_values: bool) -> None:
    clear_merges(dst_ws)
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    copy_dimensions(src_ws, dst_ws, max_row, max_col)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            copy_cell(src_ws.cell(row, col), dst_ws.cell(row, col), copy_values)
    copy_merges(src_ws, dst_ws, max_row, max_col)


def overlay_reference_styles(src_ws, dst_ws) -> None:
    max_row = min(src_ws.max_row, dst_ws.max_row)
    max_col = min(src_ws.max_column, dst_ws.max_column)
    copy_dimensions(src_ws, dst_ws, max_row, max_col)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            src_cell = src_ws.cell(row, col)
            if src_cell.has_style:
                dst_cell = dst_ws.cell(row, col)
                if isinstance(dst_cell, MergedCell):
                    continue
                current_value = dst_cell.value
                copy_cell(src_cell, dst_cell, copy_value=False)
                dst_cell.value = current_value


def replace_sheet_refs(wb, old_name: str, new_name: str) -> None:
    quoted_old = f"'{old_name}'!"
    quoted_new = f"'{new_name}'!"
    bare_old = f"{old_name}!"
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    updated = value.replace(quoted_old, quoted_new)
                    updated = updated.replace(bare_old, quoted_new)
                    if updated != value:
                        cell.value = updated


def value(ws, coord):
    return ws[coord].value


def set_style_from(src_cell, dst_cell) -> None:
    if isinstance(dst_cell, MergedCell):
        return
    current_value = dst_cell.value
    copy_cell(src_cell, dst_cell, copy_value=False)
    dst_cell.value = current_value


def fill_dane(display_ws, source_values_ws, reference_ws) -> None:
    # Keep the labels and styles from Reference, but populate with the 332044 data.
    display_ws["K4"] = value(source_values_ws, "C2")
    display_ws["J5"] = value(source_values_ws, "L15")
    display_ws["K5"] = value(source_values_ws, "M15")
    display_ws["L5"] = value(source_values_ws, "N15")
    display_ws["J6"] = value(source_values_ws, "L16")
    display_ws["K6"] = value(source_values_ws, "M16")
    display_ws["L6"] = "[m]"

    display_ws["J8"] = "ρcc"
    display_ws["K8"] = 636619.7723675814
    display_ws["J10"] = "n (obserwacje)"
    display_ws["K10"] = 32
    display_ws["J11"] = "u (niewiadome)"
    display_ws["K11"] = 16
    display_ws["J12"] = "d (defekt)"
    display_ws["K12"] = 3

    for src_row, dst_row in zip(range(5, 37), range(6, 38)):
        for src_col, dst_col in zip(range(2, 6), range(1, 5)):
            display_ws.cell(dst_row, dst_col).value = source_values_ws.cell(src_row, src_col).value
        for src_col, dst_col in zip(range(6, 10), range(5, 9)):
            display_ws.cell(dst_row, dst_col).value = source_values_ws.cell(src_row, src_col).value

    for src_row, dst_row in zip(range(5, 14), range(18, 27)):
        for src_col, dst_col in zip(range(12, 15), range(10, 13)):
            display_ws.cell(dst_row, dst_col).value = source_values_ws.cell(src_row, src_col).value

    # Reapply the reference styles on the populated areas so copied values do not disturb formats.
    for row in range(1, 38):
        for col in range(1, 19):
            set_style_from(reference_ws.cell(row, col), display_ws.cell(row, col))


def parse_base_ids(raw) -> list[int]:
    if raw is None:
        return [5, 6, 9]
    return [int(part) for part in re.findall(r"\d+", str(raw))] or [5, 6, 9]


def condition_yes_no(raw) -> str:
    return "TAK" if str(raw).strip().upper() == "SPEŁNIA" else "NIE"


def condition_good_bad(raw) -> str:
    return "dobrze" if str(raw).strip().upper() == "NIEISTOTNE" else "źle"


def fill_przemieszczenia(display_ws, source_values_ws, source_wb_values, reference_ws) -> None:
    # Top coordinate block.
    for src_row, dst_row in zip(range(14, 21), range(5, 12)):
        for src_col, dst_col in zip(range(8, 13), range(2, 7)):
            display_ws.cell(dst_row, dst_col).value = source_values_ws.cell(src_row, src_col).value

    # Summary panels.
    display_ws["C17"] = source_wb_values["Baza - wyjściowy"]["AB131"].value
    display_ws["C18"] = source_wb_values["Baza - aktualny"]["AB131"].value
    display_ws["E18"] = 2.5
    base_ids = parse_base_ids(source_values_ws["O1"].value)
    for row, base_id in zip(range(16, 19), base_ids):
        display_ws.cell(row, 7).value = base_id

    # Displacement significance table, converted to the wording used by Reference.
    good_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    bad_fill = PatternFill(fill_type="solid", fgColor="F4B6C2")
    good_font = Font(color="006100", bold=True)
    bad_font = Font(color="9C0006", bold=True)

    for src_row, dst_row in zip(range(3, 10), range(22, 29)):
        display_ws.cell(dst_row, 2).value = source_values_ws.cell(src_row, 8).value
        display_ws.cell(dst_row, 3).value = source_values_ws.cell(src_row, 9).value
        display_ws.cell(dst_row, 4).value = source_values_ws.cell(src_row, 10).value

        x_ok = condition_yes_no(source_values_ws.cell(src_row, 11).value)
        display_ws.cell(dst_row, 5).value = x_ok

        display_ws.cell(dst_row, 6).value = source_values_ws.cell(src_row, 12).value
        display_ws.cell(dst_row, 7).value = source_values_ws.cell(src_row, 13).value

        y_ok = condition_yes_no(source_values_ws.cell(src_row, 14).value)
        display_ws.cell(dst_row, 8).value = y_ok

        final_ok = condition_good_bad(source_values_ws.cell(src_row, 15).value)
        display_ws.cell(dst_row, 9).value = final_ok

    # Detailed coordinate uncertainty table.
    for point_index in range(9):
        src_start = 4 + point_index * 2
        dst_start = 34 + point_index * 2
        for offset in range(2):
            src_row = src_start + offset
            dst_row = dst_start + offset
            if offset == 0:
                display_ws.cell(dst_row, 2).value = source_values_ws.cell(src_row, 2).value
            for src_col, dst_col in zip(range(3, 7), range(3, 7)):
                display_ws.cell(dst_row, dst_col).value = source_values_ws.cell(src_row, src_col).value

    # Restore styles/number formats in populated ranges.
    for row in range(1, 54):
        for col in range(1, 12):
            set_style_from(reference_ws.cell(row, col), display_ws.cell(row, col))

    for row in range(22, 29):
        for col in (5, 8):
            cell = display_ws.cell(row, col)
            ok = cell.value == "TAK"
            cell.fill = copy(good_fill if ok else bad_fill)
            cell.font = copy(good_font if ok else bad_font)
        cell = display_ws.cell(row, 9)
        ok = cell.value == "dobrze"
        cell.fill = copy(good_fill if ok else bad_fill)
        cell.font = copy(good_font if ok else bad_font)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(SOURCE_PATH)
    source_values_wb = openpyxl.load_workbook(SOURCE_PATH, data_only=True)
    reference_wb = openpyxl.load_workbook(REFERENCE_PATH)
    if getattr(reference_wb, "loaded_theme", None):
        wb.loaded_theme = reference_wb.loaded_theme

    # Keep the original source sheets as calculation/audit sheets when replacing the display sheets.
    wb["Dane"].title = "_calc_Dane"
    wb["Przemieszczenia ostateczne"].title = "_calc_Przemieszczenia"
    replace_sheet_refs(wb, "Dane", "_calc_Dane")

    # Overlay reference styling onto calculation sheets whose tables are largely aligned.
    style_sheet_map = {
        "Wyrównanie wstępne - wyjściowy": "Wstepnewyjsciowy",
        "Wyrównanie wstępne - aktualny": "Wstepneaktualny",
        "Baza - wyjściowy": "Baza_wyj",
        "Baza - aktualny": "Baza_akt",
        "Identyfikacja": "IdentyfikacjaBazy",
    }
    for dst_name, ref_name in style_sheet_map.items():
        overlay_reference_styles(reference_wb[ref_name], wb[dst_name])

    # Build the visible Dane sheet in the Reference layout.
    dane_ws = wb.create_sheet("Dane", 0)
    copy_sheet_area(reference_wb["Dane"], dane_ws, 37, 18, copy_values=True)
    fill_dane(dane_ws, source_values_wb["Dane"], reference_wb["Dane"])

    # Build the visible Przemieszczenia sheet in the Reference layout.
    prz_index = wb.sheetnames.index("_calc_Przemieszczenia")
    prz_ws = wb.create_sheet("Przemieszczenia", prz_index)
    copy_sheet_area(reference_wb["Przemieszczenia "], prz_ws, 53, 11, copy_values=True)
    fill_przemieszczenia(
        prz_ws,
        source_values_wb["Przemieszczenia ostateczne"],
        source_values_wb,
        reference_wb["Przemieszczenia "],
    )

    wb["_calc_Dane"].sheet_state = "hidden"
    wb["_calc_Przemieszczenia"].sheet_state = "hidden"

    # Match the clean visible ordering while keeping hidden calculation sheets at the end.
    for hidden_name in ["_calc_Dane", "_calc_Przemieszczenia"]:
        ws = wb[hidden_name]
        wb._sheets.remove(ws)
        wb._sheets.append(ws)

    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
