import openpyxl

reference_path = "/Users/kjw1/Desktop/Monitoring/Proj3/Reference.xlsx"
output_path = "/Users/kjw1/Documents/New project/outputs/proj3_style_20260612/332044_proj3_popr_styl_reference_uzupelniony.xlsx"

sheet_map = {
    "Dane": "Dane",
    "Wstepnewyjsciowy": "Wstepnewyjsciowy",
    "Wstepneaktualny": "Wstepneaktualny",
    "IdentyfikacjaBazy": "IdentyfikacjaBazy",
    "Baza_wyj": "Baza_wyj",
    "Baza_akt": "Baza_akt",
    "Przemieszczenia ": "Przemieszczenia",
}

ref_wb = openpyxl.load_workbook(reference_path, data_only=False)
out_wb = openpyxl.load_workbook(output_path, data_only=False)

for ref_name, out_name in sheet_map.items():
    ref_ws = ref_wb[ref_name]
    out_ws = out_wb[out_name]
    blanks = []
    total = 0
    for row in range(1, ref_ws.max_row + 1):
        for col in range(1, ref_ws.max_column + 1):
            ref_val = ref_ws.cell(row, col).value
            if ref_val is None:
                continue
            total += 1
            out_val = out_ws.cell(row, col).value
            if out_val is None:
                blanks.append((ref_ws.cell(row, col).coordinate, ref_val))
    print(f"\n--- {ref_name} -> {out_name} ---")
    print(f"reference_nonblank={total} missing_in_output={len(blanks)} ref_range={ref_ws.max_row}x{ref_ws.max_column} out_range={out_ws.max_row}x{out_ws.max_column}")
    for coord, value in blanks[:80]:
        print(coord, repr(value)[:140])
