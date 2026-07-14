from __future__ import annotations

from copy import copy, deepcopy
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.formula import ArrayFormula

REFERENCE_PATH = Path("/Users/kjw1/Desktop/Monitoring/Proj3/Reference.xlsx")
INPUT_PATH = Path("/Users/kjw1/Documents/New project/outputs/proj3_style_20260612/332044_proj3_popr_styl_reference.xlsx")
OUTPUT_PATH = Path("/Users/kjw1/Documents/New project/outputs/proj3_style_20260612/332044_proj3_popr_styl_reference_uzupelniony.xlsx")

SHEET_MAP = {
    "Dane": "Dane",
    "Wstepnewyjsciowy": "Wstepnewyjsciowy",
    "Wstepneaktualny": "Wstepneaktualny",
    "IdentyfikacjaBazy": "IdentyfikacjaBazy",
    "Baza_wyj": "Baza_wyj",
    "Baza_akt": "Baza_akt",
    "Przemieszczenia ": "Przemieszczenia",
}

FORMULA_SHEET_MAP = {
    "Wyrównanie wstępne - wyjściowy": "Wstepnewyjsciowy",
    "Wyrównanie wstępne - aktualny": "Wstepneaktualny",
    "Identyfikacja": "IdentyfikacjaBazy",
    "Baza - wyjściowy": "Baza_wyj",
    "Baza - aktualny": "Baza_akt",
    "Przemieszczenia ": "Przemieszczenia",
}

RENAME_TO_REFERENCE = {
    "Wyrównanie wstępne - wyjściowy": "Wstepnewyjsciowy",
    "Wyrównanie wstępne - aktualny": "Wstepneaktualny",
    "Identyfikacja": "IdentyfikacjaBazy",
    "Baza - wyjściowy": "Baza_wyj",
    "Baza - aktualny": "Baza_akt",
}


def quote_sheet(name: str) -> str:
    escaped = name.replace("'", "''")
    return f"'{escaped}'"


def translate_formula_text(formula: str) -> str:
    result = formula
    for old_name, new_name in FORMULA_SHEET_MAP.items():
        result = result.replace(f"'{old_name}'!", f"{quote_sheet(new_name)}!")
        result = result.replace(f"{old_name}!", f"{quote_sheet(new_name)}!")
    return result


def translate_value(value):
    if isinstance(value, str) and value.startswith("="):
        return translate_formula_text(value)
    if isinstance(value, ArrayFormula):
        copied = deepcopy(value)
        # OpenPyXL stores the formula text on the text attribute for array formulas.
        if isinstance(copied.text, str):
            copied.text = translate_formula_text(copied.text)
        return copied
    return value


def normalize_workbook_formulas(wb) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    cell.value = translate_formula_text(value)
                elif isinstance(value, ArrayFormula):
                    if isinstance(value.text, str):
                        value.text = translate_formula_text(value.text)


def copy_style(src_cell, dst_cell) -> None:
    if isinstance(dst_cell, MergedCell):
        return
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
    dst_cell.number_format = src_cell.number_format
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.protection = copy(src_cell.protection)


def copy_dimensions(src_ws, dst_ws) -> None:
    for row_idx in range(1, src_ws.max_row + 1):
        src_dim = src_ws.row_dimensions[row_idx]
        dst_dim = dst_ws.row_dimensions[row_idx]
        if src_dim.height is not None:
            dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden
    for col_key, src_dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[col_key]
        if src_dim.width is not None:
            dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden


def clear_merges(ws) -> None:
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))


def copy_reference_merges(src_ws, dst_ws) -> None:
    for rng in src_ws.merged_cells.ranges:
        text = str(rng)
        dst_ws.merge_cells(text)


def main() -> None:
    ref_wb = openpyxl.load_workbook(REFERENCE_PATH, data_only=False)
    wb = openpyxl.load_workbook(INPUT_PATH, data_only=False)

    if getattr(ref_wb, "loaded_theme", None):
        wb.loaded_theme = ref_wb.loaded_theme

    for old_name, new_name in RENAME_TO_REFERENCE.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name
    normalize_workbook_formulas(wb)

    totals = {}
    for ref_name, out_name in SHEET_MAP.items():
        ref_ws = ref_wb[ref_name]
        out_ws = wb[out_name]
        clear_merges(out_ws)
        copy_dimensions(ref_ws, out_ws)

        filled = 0
        styled = 0
        for row in range(1, ref_ws.max_row + 1):
            for col in range(1, ref_ws.max_column + 1):
                src_cell = ref_ws.cell(row, col)
                dst_cell = out_ws.cell(row, col)
                if isinstance(dst_cell, MergedCell):
                    continue

                # Always bring over the reference style for the comparable visual area.
                copy_style(src_cell, dst_cell)
                styled += 1

                if dst_cell.value is None and src_cell.value is not None:
                    dst_cell.value = translate_value(src_cell.value)
                    filled += 1

        copy_reference_merges(ref_ws, out_ws)
        out_ws.sheet_view.showGridLines = ref_ws.sheet_view.showGridLines
        totals[out_name] = filled

    normalize_workbook_formulas(wb)

    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"

    wb.save(OUTPUT_PATH)
    for sheet, count in totals.items():
        print(f"{sheet}: filled={count}")
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
